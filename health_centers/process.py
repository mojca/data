#!/usr/bin/env python
import sys
import pathlib
# get repo root folder and add it to PYTHONPATH (so that we don't need to install this package)
repo_root_path = str(pathlib.Path(__file__).parent.absolute().parent)
if repo_root_path not in sys.path:
    sys.path.append(repo_root_path)

import collections
import csv
import datetime
import logging
import os
import re
import time
import typing
import unicodedata

import openpyxl

import health_centers.dataclass
import health_centers.get_files
import health_centers.mappings
import health_centers.utils


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__file__)


def get_sheet_hos(xlsx_file: str):
    """ We have to fetch appropriate sheet for HOS, since
            * xlsx contains multiple sheets
            * there's no guarantee relevant sheet is actually there
    """
    wb = openpyxl.load_workbook(xlsx_file)
    for sheet_name in ['Bolnišnice COVID točke', 'Bonišnice COVID točke', 'Bolnišnica COVID točke']:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
    logger.debug(f'{xlsx_file} has no relevant sheet present')


def read_sheets(sheets: typing.List[openpyxl.worksheet.worksheet.Worksheet]):
    logger.info('Standardization: Removing starting column "Št." if present...')
    for sheet in sheets:
        headers = [cell.value for cell in sheet[1]]
        if headers[0] == 'Št.':
            sheet.delete_cols(0)

    logger.info('Validating columns...')
    # file content validation
    for sheet in sheets:
        expected = [  # header row 1
            r'ZD|Bolnišnica|Bolnišnice|SB|Zdravstvena ustanova', r'Datum|datum', r'(1 )*Št. (vseh )*pregledov NMP',
            r'(2 )*Št. pregledov.* suma na COVID', r'(3 )*Št. sumov na COVID brez pregleda \(triaža po telefonu\)',
            r'(4 )*Št. opravljenih testiranj COVID', r'(5. )*Št..* pozitivnih (testov na )*COVID',
            r'(6 )*Št. napotitev (pacientov s sumom ali potrjenim COVID )*v bolnišnico',
            r'(7 )*Št. napotitev (s sumom ali potrjenim COVID )*v samoosamitev', r'Opombe|Opomba'
        ]
        actual = [cell.value for cell in sheet[1]]
        for expected_col, actual_col in zip(expected, actual):
            assert re.match(expected_col, actual_col), (sheet.file, sheet, expected_col, actual_col)

    logger.info('Reading sheets and building entity collection...')
    entities = []
    for sheet in sheets:

        for idx, row in enumerate(list(sheet.iter_rows())):  # skip header rows
            if idx == 0:  # skip header rows
                continue
            if sheet.row_dimensions[idx+1].hidden:  # indices are 1-based when querying row_dimensions
                continue
            if [cell.value for cell in row][:4] == [None, 1, 2, 3]:  # also header
                continue
            if (
                row[0].value is None and row[1].value is None or
                row[0].value == 'SKUPAJ' or
                all([cell.value == '' or cell.value is None for cell in row]) or
                any([isinstance(cell.value, str) and '=SUBTOTAL(' in cell.value for cell in row])
            ):
                break  # aggregates do not interest us, also do not continue since all the relevant data is extracted

            # make sure the name is always there
            assert row[0].value is not None, (sheet.file, sheet, f'row index: {idx}', [cell.value for cell in row])
            # make sure the date is always there
            if row[1].value is None:  # some cells are missing dates
                search = re.search(r'<Worksheet "(\d{1,2})\.(\d{1,2})\.">', str(sheet))
                assert datetime.datetime.now().year == 2020  # this extraction logic could be broken in 2021
                date = datetime.datetime(year=2020, month=int(search.group(2)), day=int(search.group(1))).date()
            else:
                date = row[1].value.date()
            try:
                entity = health_centers.dataclass.Entity(
                    name=row[0].value,
                    # the date we get from xslx means date when data was gathered
                    # the data itself is for the day before that
                    date=date - datetime.timedelta(days=1),
                    sheet=str(sheet),
                    file=sheet.file,
                    numbers=health_centers.dataclass.Numbers(
                        # Št. pregledov NMP
                        examinations___medical_emergency=row[2].value,
                        # Št. pregledov  suma na COVID
                        examinations___suspected_covid=row[3].value,
                        # Št. sumov na COVID brez pregleda (triaža po telefonu)
                        phone_triage___suspected_covid=row[4].value,
                        # Št. opravljenih testiranj COVID
                        tests___performed=row[5].value,
                        # Št.  pozitivnih COVID
                        tests___positive=row[6].value,
                        # Št. napotitev v bolnišnico
                        sent_to___hospital=row[7].value,
                        # Št. napotitev v samoosamitev
                        sent_to___self_isolation=row[8].value
                    )
                )
            except Exception as e:
                raise type(e)(str(e), sheet.file)
            entities.append(entity)
    return entities


@health_centers.utils.timeit
def get_sheets_hos(files: typing.List[str]):
    sheets = []
    for f in files:
        sheet = get_sheet_hos(xlsx_file=f)
        if sheet is not None:
            sheet.file = f
            sheets.append(sheet)
    return sheets


@health_centers.utils.timeit
def get_sheets_zd(files: typing.List[str]):
    sheets = []
    for f in files:
        sheet = openpyxl.load_workbook(f).active
        sheet.file = f
        sheets.append(sheet)
    return sheets


@health_centers.utils.timeit
def main():

    cache = health_centers.utils.get_cache()
    files = health_centers.get_files.main()
    for key in list(cache):  # clear the cache of non-relevant records
        if key not in files.all:
            del cache[key]

    to_be_processed_files_hos = [f for f in files.hos if f not in cache]
    to_be_processed_files_zd = [f for f in files.zd if f not in cache]
    for f in to_be_processed_files_hos + to_be_processed_files_zd:
        cache[f] = []  # pre-populate so that we also have filenames with zero information cached

    fresh_entities_hos = read_sheets(sheets=get_sheets_hos(files=to_be_processed_files_hos))
    fresh_entities_zd = read_sheets(sheets=get_sheets_zd(files=to_be_processed_files_zd))
    for e in fresh_entities_hos + fresh_entities_zd:
        cache[e.file].append(e)
    health_centers.utils.set_cache(cache)

    entities = [e for entities in cache.values() for e in entities]
    entities.sort(key=lambda entity: entity.name)
    entities.sort(key=lambda entity: entity.date)

    aggregates = collections.defaultdict(lambda: health_centers.dataclass.Numbers(0, 0, 0, 0, 0, 0, 0))
    for entity in entities:
        for key in health_centers.dataclass.Numbers.__annotations__.keys():
            aggregates[entity.date].__dict__[key] += entity.numbers.__dict__[key] or 0  # handle Null

    logger.info('Writing CSV...')
    repo_root_health_centers = os.path.dirname(os.path.abspath(__file__))
    assert repo_root_health_centers.endswith('/data/health_centers')
    repo_root = '/'.join(repo_root_health_centers.split('/')[:-1])
    assert repo_root.endswith('/data')
    health_centers_csv = os.path.join(repo_root, 'csv/health_centers.csv')

    def get_entity(name_key: str, date: datetime):
        found_entities = []
        for entity in entities:
            if entity.name_key == name_key and entity.date == date:
                found_entities.append(entity)
        if len(found_entities) == 0:
            logger.debug(f'No data found for {name_key} {date}')
            return None
        if len(found_entities) > 1:

            # it might happen that numbers come from diffent files, but if they are the same that's okay
            if len(set([e.numbers for e in found_entities])) == 1:
                return found_entities[0]

            # if numbers are not the same, then we take the maximum over all properties (not sum)
            # this is not ideal scenario though, and should maybe be removed in the future
            props = set([  # this definition is here so that we are sure maximums make sense for all the fields
                'examinations___medical_emergency', 'examinations___suspected_covid',
                'phone_triage___suspected_covid', 'tests___performed', 'tests___positive', 'sent_to___hospital',
                'sent_to___self_isolation'
            ])
            maxs = {p: 0 for p in props}
            for e in found_entities:
                assert e.numbers.__annotations__.keys() == props
                for a in props:
                    if (e.numbers.get(a) or -1) > maxs[a]:  # property can be None, therefore; or 0
                        maxs[a] = e.numbers.get(a)
            for e in found_entities:
                for prop in props:
                    if e.numbers.get(prop) == maxs[prop]:
                        return e

            # code unreachable (for now)
            for found_entity in found_entities:
                logger.error(found_entity)
            raise Exception(f'Too many entities found: {len(found_entities)}, {name_key}, {date}')
        return found_entities[0]

    with open(health_centers_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, dialect='excel', lineterminator='\n')

        def get_formatted_numbers_fields():
            return[field.replace('___', '.') for field in health_centers.dataclass.Numbers.__annotations__.keys()]

        columns = ['date']
        # scope: aggregates
        for field in get_formatted_numbers_fields():
            columns.append(f'hc.{field}')
        # scope: health centers
        for name in health_centers.mappings.unique_short_names:
            region = health_centers.mappings.region[name]
            for field in get_formatted_numbers_fields():
                columns.append(f'hc.{region}.{name}.{field}')
        writer.writerow(columns)

        # write data
        dates = sorted(list(set([entity.date for entity in entities])))
        for date in dates:

            columns = [date]
            # scope: aggregates
            for key in aggregates[date].__annotations__.keys():
                columns.append(aggregates[date].__dict__[key])
            # scope: health centers
            for name in health_centers.mappings.unique_short_names:
                entity = get_entity(name_key=name, date=date)
                for field in health_centers.dataclass.Numbers.__annotations__.keys():
                    if entity is None:
                        columns.append(None)
                    else:
                        columns.append(getattr(entity.numbers, field))
            writer.writerow(columns)

    logger.info('Writing CSV timestamp...')
    with open(f'{health_centers_csv}.timestamp', 'w') as timestamp_file:
        timestamp = int(time.time())
        timestamp_file.write(f'{timestamp}\n')

    logger.info('Writing one dimensional output for easier checking/diff purposes...')
    check_csv = os.path.join(repo_root, 'csv/health_centers_check.csv')
    with open(check_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, dialect='excel', lineterminator='\n')
        for entity in entities:  # should be already sorted
            writer.writerow([
                entity.date,
                entity.name,
                entity.name_key,
                entity.sheet,
                # to ensure standardized output for different operating systems we normalize filenames
                unicodedata.normalize('NFC', entity.file.split('/')[-1]),
                entity.numbers.examinations___medical_emergency,
                entity.numbers.examinations___suspected_covid,
                entity.numbers.phone_triage___suspected_covid,
                entity.numbers.tests___performed,
                entity.numbers.tests___positive,
                entity.numbers.sent_to___hospital,
                entity.numbers.sent_to___self_isolation
            ])


if __name__ == '__main__':
    logger.info('Starting...')
    main()
